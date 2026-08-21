import sqlite3
import pandas as pd
from datetime import datetime, timedelta
import logging
import re, os, glob, csv, json
import yaml as _yaml
import shutil
from typing import List, Dict, Any, Set, Optional

logger = logging.getLogger(__name__)

# [AI-2026-08-15] 套利基金代码白名单缓存（读 arbcore/config/lof_config.yaml，mtime 感知热加载）
_LOF_CFG_PATH = os.path.normpath(os.path.join(os.path.dirname(__file__), '..', '..', '..', 'arbcore', 'config', 'lof_config.yaml'))
_ARB_CODES_CACHE: Dict[str, Any] = {"codes": None, "mtime": 0}

def _get_arbitrage_codes() -> Set[str]:
    """从 lof_config.yaml 读所有套利基金代码（funds[].code）作为白名单。"""
    try:
        mtime = os.path.getmtime(_LOF_CFG_PATH)
    except OSError:
        return set()
    if _ARB_CODES_CACHE["codes"] is not None and _ARB_CODES_CACHE["mtime"] == mtime:
        return _ARB_CODES_CACHE["codes"]
    try:
        with open(_LOF_CFG_PATH, 'r', encoding='utf-8') as f:
            cfg = _yaml.safe_load(f)
        codes = {str(x.get('code', '')).strip() for x in cfg.get('funds', [])}
        codes.discard('')
        _ARB_CODES_CACHE["codes"] = codes
        _ARB_CODES_CACHE["mtime"] = mtime
        logger.info(f"[TDX-PreFilter] 套利白名单载入: {len(codes)} 只 (lof_config.yaml)")
        return codes
    except Exception as e:
        logger.warning(f"[TDX-PreFilter] 加载 lof_config.yaml 失败: {e}")
        return set()

class LedgerService:
    def __init__(self, db_manager, master_db_path=None):
        self.db = db_manager  # 交易库（arb_tran.db）
        # master 库路径（市场因子：fund_daily_factors.hedge / exchange_rate），对账时 ATTACH 只读
        self.master_db_path = master_db_path

    def get_all_trades(self, status: str = 'ACTIVE') -> List[Dict[str, Any]]:
        """获取所有实盘记录"""
        conn = self.db._get_conn()
        try:
            query = "SELECT * FROM user_trades WHERE status = ? ORDER BY trade_date DESC"
            df = pd.read_sql_query(query, conn, params=(status,))
            
            # 增强逻辑：计算剩余赎回天数与染色状态
            today = datetime.now().date()
            trades = df.to_dict(orient='records')
            for t in trades:
                if t['remind_date']:
                    remind = datetime.strptime(t['remind_date'], '%Y-%m-%d').date()
                    t['days_left'] = (remind - today).days
                else:
                    t['days_left'] = None
            return trades
        finally:
            conn.close()

    def _get_next_workday(self, current_date: datetime, days: int) -> datetime:
        """计算 N 个交易日后的日期 (跳过周六日)"""
        added_days = 0
        tmp_date = current_date
        while added_days < days:
            tmp_date += timedelta(days=1)
            if tmp_date.weekday() < 5: # 0-4 是周一到周五
                added_days += 1
        return tmp_date

    def add_trade(self, trade_data: Dict[str, Any]):
        """
        新增对账记录
        """
        conn = self.db._get_conn()
        try:
            trade_date_str = trade_data.get('trade_date', datetime.now().strftime('%Y-%m-%d'))
            dt = datetime.strptime(trade_date_str, '%Y-%m-%d')
            
            # [V4.6 核心规则]：自动推演 3 个交易日后的赎回日
            # 如果前端传了手动修改后的 remind_date，优先使用手动值
            manual_remind = trade_data.get('remind_date')
            if manual_remind and manual_remind != '':
                remind_date = manual_remind
            else:
                # 否则执行自动推演逻辑 (T+3 工作日)
                remind_dt = self._get_next_workday(dt, 3)
                remind_date = remind_dt.strftime('%Y-%m-%d')
            
            query = """
                INSERT INTO user_trades 
                (fund_code, fund_name, account_suffix, action, volume, price, amount, 
                 hedge_symbol, hedge_price, hedge_vol, fees, trade_date, remind_date, status)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'ACTIVE')
            """
            conn.execute(query, (
                trade_data['fund_code'],
                trade_data.get('fund_name', ''),
                trade_data.get('account_suffix', ''),
                trade_data['action'],
                trade_data['volume'],
                trade_data['price'],
                float(trade_data['volume']) * float(trade_data['price']),
                trade_data.get('hedge_symbol'),
                trade_data.get('hedge_price'),
                trade_data.get('hedge_vol'),
                trade_data.get('fees', 0),
                trade_date_str,
                remind_date
            ))
            conn.commit()
            return True
        except Exception as e:
            logger.error(f"记账失败: {e}")
            return False
        finally:
            conn.close()

    def close_trade(self, trade_id: int):
        conn = self.db._get_conn()
        try:
            conn.execute("UPDATE user_trades SET status = 'CLOSED' WHERE id = ?", (trade_id,))
            conn.commit()
            return True
        finally:
            conn.close()

    # --- 费率管理 ---
    def get_fund_fees(self, fund_code: str) -> Dict[str, Any]:
        conn = self.db._get_conn()
        try:
            df = pd.read_sql_query("SELECT * FROM fund_fees WHERE fund_code = ?", conn, params=(fund_code,))
            if not df.empty:
                return df.iloc[0].to_dict()
            return {"redemption_fee_rate": 0.5, "broker_name": ""}
        finally:
            conn.close()

    def upsert_fund_fee(self, data: Dict[str, Any]):
        conn = self.db._get_conn()
        try:
            query = "INSERT OR REPLACE INTO fund_fees (fund_code, redemption_fee_rate, broker_name, updated_at) VALUES (?, ?, ?, datetime('now'))"
            conn.execute(query, (data['fund_code'], data['redemption_fee_rate'], data.get('broker_name', '')))
            conn.commit()
            return True
        except Exception as e:
            logger.error(f"upsert_fund_fee failed: {e}")
            return False
        finally:
            conn.close()

    def get_broker_redemption_fees(self):
        conn = self.db._get_conn()
        try:
            df = pd.read_sql_query("SELECT * FROM broker_redemption_fees", conn)
            return df.to_dict('records')
        finally:
            conn.close()

    def upsert_broker_redemption_fee(self, data: Dict[str, Any]):
        conn = self.db._get_conn()
        try:
            query = "INSERT OR REPLACE INTO broker_redemption_fees (category, fund_code, broker_name, fee_rate, updated_at) VALUES (?, ?, ?, ?, datetime('now', 'localtime'))"
            conn.execute(query, (data.get('category', ''), data['fund_code'], data['broker_name'], data['fee_rate']))
            conn.commit()
            return True
        except Exception as e:
            logger.error(f"upsert_broker_redemption_fee failed: {e}")
            return False
        finally:
            conn.close()

    def delete_broker_redemption_fee(self, fee_id: int):
        conn = self.db._get_conn()
        try:
            conn.execute("DELETE FROM broker_redemption_fees WHERE id = ?", (fee_id,))
            conn.commit()
            return True
        except Exception as e:
            logger.error(f"delete_broker_redemption_fee failed: {e}")
            return False
        finally:
            conn.close()

    # ================================================================
    # 辅助方法（默认价格、费率等）
    # ================================================================

    def get_prev_close(self, fund_code: str) -> float:
        """获取最新收盘价"""
        try:
            conn = self.db._get_conn()
            try:
                cur = conn.execute(
                    "SELECT price FROM unified_fund_history WHERE fund_code=? AND price IS NOT NULL ORDER BY date DESC LIMIT 1",
                    (fund_code,)
                )
                row = cur.fetchone()
                return float(row[0]) if row else 0
            finally:
                conn.close()
        except:
            return 0

    def get_fee_rate(self, fund_code: str, broker: str = '') -> float:
        """获取指定基金+券商的赎回费率"""
        try:
            conn = self.db._get_conn()
            try:
                if broker:
                    cur = conn.execute(
                        "SELECT fee_rate FROM broker_redemption_fees WHERE fund_code=? AND broker_name=? LIMIT 1",
                        (fund_code, broker)
                    )
                else:
                    cur = conn.execute(
                        "SELECT fee_rate FROM broker_redemption_fees WHERE fund_code=? LIMIT 1",
                        (fund_code,)
                    )
                row = cur.fetchone()
                return float(row[0]) if row else 0
            finally:
                conn.close()
        except:
            return 0

    # ================================================================
    # 套利对账本（arbitrage_pairs）- 匹配Excel格式
    # ================================================================

    def _attach_master(self, conn) -> str:
        """ATTACH 主库（arb_master.db，含每日 exchange_rate）为只读别名 master。

        [AI-2026-08-18] 此前该方法从未定义却被 _get_usd_rate/_get_rates_map 调用，
        AttributeError 被吞 → 汇率永远回退 7.2，所有 pnl_rmb 系统性算错。现补实现。
        每次调用先 DETACH（忽略不存在错误）再 ATTACH，保证幂等。
        """
        if not self.master_db_path:
            raise RuntimeError("master_db_path 未配置，无法读取主库汇率")
        try:
            conn.execute("DETACH DATABASE master")
        except Exception:
            pass
        conn.execute("ATTACH DATABASE ? AS master", (self.master_db_path,))
        return 'master'

    def _get_usd_rate(self) -> Optional[float]:
        """获取最新美元汇率（主库 exchange_rate 最后一天）。无数据返回 None，不兜底（SUPREME 铁律）。"""
        try:
            conn = self.db._get_conn()
            try:
                m = self._attach_master(conn)
                cur = conn.execute(
                    f"SELECT usd_cny_mid FROM {m}.exchange_rate ORDER BY date DESC LIMIT 1"
                )
                row = cur.fetchone()
                return float(row[0]) if row else None
            finally:
                conn.close()
        except Exception:
            return None

    def _get_rates_map(self, conn) -> dict:
        """一次性加载 exchange_rate 全表 -> {date: usd_cny_mid}（升序），供按日取汇率"""
        try:
            m = self._attach_master(conn)
            rows = conn.execute(
                f"SELECT date, usd_cny_mid FROM {m}.exchange_rate WHERE usd_cny_mid IS NOT NULL ORDER BY date"
            ).fetchall()
            return {r[0]: float(r[1]) for r in rows}
        except Exception:
            return {}

    @staticmethod
    def _rate_on(rates: dict, date_str) -> Optional[float]:
        """取指定日期（或之前最近一天）的美元中间价；无数据返回 None（不兜底，符合 SUPREME 铁律）。

        [AI-2026-08-18] 原实现无数据回退 7.2 → 所有组 pnl_rmb 系统性算错。现改为缺失即 None，
        由调用方（import_v7）显式提示并留空，绝不伪装汇率。
        """
        if not rates:
            return None
        if date_str:
            if date_str in rates:
                return rates[date_str]
            prior = [k for k in rates if k <= date_str]
            if prior:
                return rates[max(prior)]
            return None  # date 早于汇率表最早一天
        return None  # 无日期：不猜（瘸腿组应补日期后重导）

    def get_all_pairs(self, status: str = None) -> List[Dict[str, Any]]:
        """获取套利对列表"""
        conn = self.db._get_conn()
        try:
            self._ensure_a_share_pnl(conn)
            if status:
                df = pd.read_sql_query(
                    "SELECT * FROM arbitrage_pairs WHERE status = ? ORDER BY buy_date DESC, serial_no DESC",
                    conn, params=(status,)
                )
            else:
                df = pd.read_sql_query(
                    "SELECT * FROM arbitrage_pairs ORDER BY buy_date DESC, serial_no DESC",
                    conn
                )
            pairs = df.to_dict(orient='records')
            # [AI-2026-08-16] pandas 把 SQL NULL 读成 NaN，NaN 无法 JSON 序列化（会 500），转回 None
            for p in pairs:
                for k, v in list(p.items()):
                    if isinstance(v, float) and (v != v or v in (float('inf'), float('-inf'))):
                        p[k] = None
            # [AI-2026-08-18] a_share_pnl/pnl_usd 由 V7 导入时按明细 G/S 求和落库，查询直接读列；不再动态反算
            for p in pairs:
                p['us_pnl'] = p.get('pnl_usd')
            return pairs
        finally:
            conn.close()

    def add_pair(self, data: Dict[str, Any]) -> int:
        """新增套利对"""
        conn = self.db._get_conn()
        try:
            self._ensure_open_type(conn)
            buy_vol = data.get('buy_volume') or 0
            buy_price = data.get('buy_price') or 0
            buy_amount = data.get('buy_amount') or (buy_vol * buy_price)
            short_vol = data.get('short_volume') or 0
            short_price = data.get('short_price') or 0
            short_amount = data.get('short_amount') or (short_vol * short_price)

            rates = self._get_rates_map(conn)
            rate = self._rate_on(rates, data.get('sell_date') or data.get('buy_date'))
            sell_amt = data.get('sell_amount') or 0
            redeem_fee = data.get('redemption_fee') or 0
            cover_amt = data.get('cover_amount') or 0
            us_comm = data.get('us_commission') or 0

            a_pnl = sell_amt - buy_amount - redeem_fee
            u_pnl = (cover_amt - short_amount) - us_comm
            pnl_rmb = round(a_pnl + u_pnl * rate, 2)
            pnl_usd = round(u_pnl, 2)
            a_share_pnl = round(a_pnl, 2) if (data.get('status') == 'Closed') else None

            conn.execute('''
                INSERT INTO arbitrage_pairs
                (fund_code, fund_name, buy_date, buy_price, buy_volume, buy_amount, buy_account,
                 sell_date, sell_price, sell_volume, sell_amount, redemption_fee,
                 hedge_symbol, short_date, short_price, short_volume, short_amount,
                 cover_date, cover_price, cover_volume, cover_amount, us_commission,
                 pnl_rmb, pnl_usd, status, buy_notes, sell_notes, notes,
                 broker_name, open_type, close_type, a_share_pnl)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            ''', (
                data['fund_code'], data.get('fund_name', ''),
                data.get('buy_date'), buy_price, buy_vol, buy_amount, data.get('buy_account'),
                data.get('sell_date'), data.get('sell_price'), data.get('sell_volume'), sell_amt, redeem_fee,
                data.get('hedge_symbol'), data.get('short_date'), short_price, short_vol, short_amount,
                data.get('cover_date'), data.get('cover_price'), data.get('cover_volume'), cover_amt, us_comm,
                pnl_rmb, pnl_usd, data.get('status', 'ACTIVE'),
                data.get('buy_notes'), data.get('sell_notes'), data.get('notes'),
                data.get('broker_name', ''), data.get('open_type', 'BUY'), data.get('close_type', 'REDEEM'),
                a_share_pnl
            ))
            conn.commit()
            return conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        except Exception as e:
            logger.error(f"新增套利对失败: {e}")
            raise
        finally:
            conn.close()

    def update_pair(self, pair_id: int, data: Dict[str, Any]) -> bool:
        """更新套利对"""
        conn = self.db._get_conn()
        try:
            self._ensure_a_share_pnl(conn)
            fields = []
            values = []
            for key in ['fund_code','fund_name','buy_date','buy_price','buy_volume','buy_amount',
                        'buy_account','sell_date','sell_price','sell_amount','redemption_fee',
                        'hedge_symbol','short_date','short_price','short_volume','short_amount',
                        'cover_date','cover_price','cover_amount','us_commission',
                        'status','buy_notes','sell_notes','notes',
                        'broker_name','open_type','close_type']:
                if key in data:
                    fields.append(f"{key} = ?")
                    values.append(data[key])

            if not fields:
                return False

            # 盈亏重算守卫：仅 Closed 才重算/保留 pnl；OPEN/unfinished 一律置 NULL
            # 口径：pnl 只在 status='Closed' 时非 NULL（赎回日净值已出，真收益可算）；
            #       OPEN/unfinished 既算不出、又易因 sell_amount=0 被公式误记成巨亏。
            effective_status = data.get('status')
            if effective_status is None:
                df0 = pd.read_sql_query(
                    "SELECT status FROM arbitrage_pairs WHERE id = ?", conn, params=(pair_id,)
                )
                if not df0.empty:
                    effective_status = df0.iloc[0]['status']

            if effective_status != 'Closed':
                # 非 Closed：不重算，强制 pnl 置 NULL（与 V7 汇总行收益留空对齐）
                fields.append("pnl_rmb = ?")
                fields.append("pnl_usd = ?")
                fields.append("a_share_pnl = ?")
                values.extend([None, None, None])
            elif any(k in data for k in ['buy_amount','sell_amount','redemption_fee',
                                         'short_amount','cover_amount','us_commission']):
                # Closed 且金额变化：正常重算盈亏
                buy_amt = data.get('buy_amount') or 0
                sell_amt = data.get('sell_amount') or 0
                redeem_fee = data.get('redemption_fee') or 0
                short_amt = data.get('short_amount') or 0
                cover_amt = data.get('cover_amount') or 0
                us_comm = data.get('us_commission') or 0

                if not all([buy_amt, sell_amt]):
                    df = pd.read_sql_query("SELECT * FROM arbitrage_pairs WHERE id = ?", conn, params=(pair_id,))
                    if not df.empty:
                        row = df.iloc[0]
                        buy_amt = buy_amt or (row.get('buy_amount') or 0)
                        sell_amt = sell_amt or (row.get('sell_amount') or 0)
                        redeem_fee = redeem_fee or (row.get('redemption_fee') or 0)
                        short_amt = short_amt or (row.get('short_amount') or 0)
                        cover_amt = cover_amt or (row.get('cover_amount') or 0)
                        us_comm = us_comm or (row.get('us_commission') or 0)

                rates = self._get_rates_map(conn)
                rate = self._rate_on(rates, data.get('sell_date') or data.get('buy_date'))
                a_pnl = sell_amt - buy_amt - redeem_fee
                u_pnl = (cover_amt - short_amt) - us_comm
                pnl_rmb = round(a_pnl + u_pnl * rate, 2)
                pnl_usd = round(u_pnl, 2)
                fields.append("pnl_rmb = ?")
                fields.append("pnl_usd = ?")
                fields.append("a_share_pnl = ?")
                values.extend([pnl_rmb, pnl_usd, round(a_pnl, 2)])

            fields.append("updated_at = datetime('now', 'localtime')")
            values.append(pair_id)
            conn.execute(f"UPDATE arbitrage_pairs SET {', '.join(fields)} WHERE id = ?", values)
            conn.commit()
            return True
        except Exception as e:
            logger.error(f"更新套利对失败: {e}")
            return False
        finally:
            conn.close()

    def delete_pair(self, pair_id: int) -> bool:
        """删除套利对"""
        conn = self.db._get_conn()
        try:
            conn.execute("DELETE FROM arbitrage_pairs WHERE id = ?", (pair_id,))
            conn.commit()
            return True
        except Exception as e:
            logger.error(f"删除套利对失败: {e}")
            return False
        finally:
            conn.close()

    def auto_record_trade(self, data: Dict[str, Any]) -> int:
        """自动记录一笔成交（从QMT交易回调）"""
        conn = self.db._get_conn()
        try:
            action = data.get('action', 'BUY')
            fund_code = data.get('fund_code', '')
            price = data.get('price', 0)
            volume = int(data.get('volume', 0))
            amount = data.get('amount', 0) or (price * volume)

            if action == 'BUY':
                # A股买入 → 新建一个套利对
                pair_id = self.add_pair({
                    'fund_code': fund_code.split('.')[0],
                    'fund_name': data.get('fund_name', ''),
                    'buy_date': data.get('trade_date', datetime.now().strftime('%Y-%m-%d')),
                    'buy_price': price,
                    'buy_volume': volume,
                    'buy_amount': amount,
                    'buy_account': data.get('account_suffix', ''),
                    'buy_notes': data.get('notes', '自动记录'),
                    'status': 'ACTIVE'
                })
                return pair_id
            elif action == 'SELL':
                # 美股做空 → 找到该基金最新没有美股侧的ACTIVE对，附加上去
                df = pd.read_sql_query(
                    "SELECT id FROM arbitrage_pairs WHERE status='ACTIVE' AND (short_amount IS NULL OR short_amount=0) AND fund_code=? ORDER BY id DESC LIMIT 1",
                    conn, params=(fund_code.split('.')[0],)
                )
                if not df.empty:
                    pair_id = int(df.iloc[0]['id'])
                    self.update_pair(pair_id, {
                        'hedge_symbol': data.get('hedge_symbol', ''),
                        'short_date': data.get('trade_date'),
                        'short_price': price,
                        'short_volume': volume,
                        'short_amount': amount
                    })
                    return pair_id
            return 0
        except Exception as e:
            logger.error(f"自动记录交易失败: {e}")
            return 0
        finally:
            conn.close()

    # ================================================================
    # [AI-2026-08-15] IB 真实成交流水（从 IB reqExecutions 同步）
    # ================================================================

    def _parse_local_date(self, trade_time: str):
        """IB 成交时间 -> 本地日期 YYYY-MM-DD（用于按天过滤）。"""
        if not trade_time:
            return None
        try:
            return datetime.strptime(trade_time, "%Y%m%d %H:%M:%S").strftime("%Y-%m-%d")
        except:
            pass
        try:
            return datetime.fromisoformat(trade_time).strftime("%Y-%m-%d")
        except:
            return None

    def _ensure_open_type(self, conn):
        """确保 arbitrage_pairs.open_type 列存在（记录 A股开仓真实类型 BUY/SUBSCRIBE），幂等。"""
        try:
            cols = [r[1] for r in conn.execute("PRAGMA table_info(arbitrage_pairs)").fetchall()]
            if "open_type" not in cols:
                conn.execute("ALTER TABLE arbitrage_pairs ADD COLUMN open_type TEXT DEFAULT 'BUY'")
                conn.commit()
        except Exception:
            pass

    def _ensure_a_share_pnl(self, conn):
        """[AI-2026-08-18] 确保 arbitrage_pairs.a_share_pnl 列存在（A股盈亏 RMB，V7 明细 G 列求和落库），幂等。"""
        try:
            cols = [r[1] for r in conn.execute("PRAGMA table_info(arbitrage_pairs)").fetchall()]
            if "a_share_pnl" not in cols:
                conn.execute("ALTER TABLE arbitrage_pairs ADD COLUMN a_share_pnl REAL")
                conn.commit()
        except Exception:
            pass


    # ================================================================
    # [AI-2026-08-16] 赎回告警：OPEN 笔推算可优惠赎回日 + unfinished 待净值
    # ================================================================
    def _discount_redeem_date(self, buy_date_str: str):
        """按买入星期推算可优惠赎回日：周一/二买+3天、周三/四/五买+5天，跳过周末。"""
        try:
            buy = datetime.strptime(buy_date_str, '%Y-%m-%d').date()
        except Exception:
            return None
        wd = buy.weekday()  # Mon=0..Sun=6
        if wd in (0, 1):
            days = 3
        elif wd in (2, 3, 4):
            days = 5
        else:  # 周六/周日（A股不会出现，兜底+7）
            days = 7
        d = buy + timedelta(days=days)
        while d.weekday() >= 5:  # 跳过周末
            d += timedelta(days=1)
        return d

    # [AI-2026-08-16] 导入 v7 Excel 套利账本：解析 -> upsert 到 arbitrage_pairs
    _HEDGE_MAP = {'162411': 'XOP', '164701': 'GLD', '161116': 'GLD', '164824': 'INDA'}
    _STATUS_MAP = {'closed': 'Closed', 'final': 'Closed', 'open': 'OPEN', 'unfinished': 'unfinished'}
    # [AI-2026-08-18] 用户习惯叫法（优先于主库官方名）：统一前端展示口径，勿改
    _HABIT_FUND_NAMES = {'161116': '易方达黄金', '162411': '华宝油气', '164701': '汇添富黄金', '164824': '印度基金'}

    def _parse_v7_groups(self, file_path: str) -> List[Dict[str, Any]]:
        """解析 v7 Excel，返回 [{summary, details}] 列表。每组配对由连续的明细行 + 一个汇总行组成。"""
        STATUS_WORDS = {'closed', 'final', 'open', 'unfinished'}

        # ---- 尝试 openpyxl（首选），失败则 fallback pandas ----
        wb = None
        ws = None
        use_pandas = False
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
        except Exception as e:
            import traceback
            logger.warning(f"openpyxl 解析 V7 失败({e})，fallback pandas: {traceback.format_exc()}")
            use_pandas = True

        if use_pandas:
            return self._parse_v7_pandas(file_path)

        def _num(v):
            if v is None:
                return None
            if isinstance(v, bool):
                return None
            if isinstance(v, (int, float)):
                return float(v)          # 数值保留真实符号, 避免负号被重复取反
            s = str(v).strip()
            neg = s.startswith('(') and s.endswith(')')
            s2 = s.replace('¥', '').replace('$', '').replace(',', '').replace('(', '-').replace(')', '')
            try:
                return float(s2)
            except Exception:
                return None

        def _date(v):
            if v is None:
                return None
            if hasattr(v, 'year'):
                return v.strftime('%Y-%m-%d')
            s = str(v).strip()
            m = re.search(r'(\d{4})[-/](\d{1,2})[-/](\d{1,2})', s)
            if m:
                return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
            return None

        groups = []
        cur = None
        for r in range(2, ws.max_row + 1):
            A = ws.cell(r, 1).value    # 序号
            B = ws.cell(r, 2).value    # 日期/状态词
            E = ws.cell(r, 5).value    # 基金
            F = ws.cell(r, 6).value    # 动作
            N = ws.cell(r, 14).value  # 收益RMB
            O = ws.cell(r, 15).value  # 对冲量
            P = ws.cell(r, 16).value  # 对冲价
            S = ws.cell(r, 19).value  # IB小计USD
            I = ws.cell(r, 9).value   # I列 单价(买入加权均价)
            K = ws.cell(r, 11).value  # K列 备注
            L = ws.cell(r, 12).value  # L列 美股标的
            Q = ws.cell(r, 17).value  # Q列 对冲额(东哥已填SUM)
            R = ws.cell(r, 18).value  # R列 佣金(东哥已填SUM)
            # [AI-2026-08-18] 空行判定须含 G/O/S：纯美股对冲行（F 空但 O/S 有值）不是空行
            Gv = ws.cell(r, 7).value
            if (A is None and B is None and E is None and F is None and N is None
                    and Gv is None and O is None and S is None):
                # [AI-2026-08-18 东哥口径] 空行只是格式占位，不参与组边界判断；
                # 组结束唯一标志 = 汇总行（F='汇总'）。空行直接跳过，绝不切断/收组。
                continue
            E = str(E).strip() if E else ''
            F = str(F).strip() if F else ''
            A_str = str(A).strip() if A else ''
            B_str = str(B).strip() if B else ''
            if F == '汇总':
                if cur is None:
                    cur = {'details': [], 'fund': E}
                cur['fund'] = E or cur.get('fund')
                # 汇总行 B 列是状态词(Closed/OPEN/unfinished)，非日期
                b_is_date = hasattr(B, 'year') or (B_str and B_str.lower() not in STATUS_WORDS)
                status = self._STATUS_MAP.get(B_str.lower(), 'Final' if b_is_date else B_str)
                G = ws.cell(r, 7).value    # G列=多账户交易金额(LOF买入总额)
                summary = {
                    'serial': A_str,
                    'status': status,
                    'fund': E,
                    'row': r,   # [AI-2026-08-18] 汇总行 Excel 行号，供导入后回填 T 列汇率/N 列公式
                    'pnl_rmb': _num(N),
                    'pnl_usd': _num(S),
                    'buy_amount': _num(G),   # V7习惯写负数(资金流出)，导入时归一为正数
                    'buy_price': _num(I),    # I列 单价(汇总行加权均价)
                    'short_price': _num(P),  # P列 对冲价(汇总行)
                    'short_amount': _num(Q),  # Q列 对冲额(汇总行, 东哥已填SUM)
                    'us_commission': _num(R),  # R列 佣金(汇总行, 东哥已填SUM)
                    'hedge_symbol': (str(L).strip() if L is not None else ''),  # L列 美股标的
                    'buy_notes': (str(K).strip() if K is not None else ''),     # K列 备注
                    'buy_date': None,
                    'sell_date': None,
                }
                cur['summary'] = summary
                groups.append(cur)
                cur = None
                continue
            has_us = (N is not None) or (O is not None) or (S is not None)
            if E or has_us:
                if cur is None:
                    cur = {'details': [], 'fund': E}
                if E:
                    cur['fund'] = E
                cur['details'].append({
                    'action': F or '美股明细',
                    'date': _date(B),
                    'volume': _num(ws.cell(r, 10).value),
                    'short_vol': _num(O),
                    'short_price': _num(P),
                    'price': _num(ws.cell(r, 9).value),  # I列 单价(买入/赎回价)
                    'comm': _num(ws.cell(r, 18).value),  # R列 佣金(东哥已填SUM)
                    'g_amt': _num(ws.cell(r, 7).value),   # G列 A股交易金额(买入负/赎回正)
                    's_amt': _num(ws.cell(r, 19).value),  # S列 IB小计USD(卖空正/买平负)
                })
        if cur:
            if 'summary' not in cur:
                logger.warning(f"V7 第 {r} 行附近（文件末尾）：存在未以'汇总'行结尾的明细块，该组已跳过（不影响其他组）")
            groups.append(cur)
        return groups

    def _parse_v7_pandas(self, file_path: str) -> List[Dict[str, Any]]:
        """pandas 兜底解析 V7 Excel（openpyxl XMLParser 报错时自动切换）。"""
        import pandas as pd
        STATUS_WORDS = {'closed', 'final', 'open', 'unfinished'}

        df = pd.read_excel(file_path, header=None)  # 0-indexed，第1行=表头

        def _num(v):
            if v is None:
                return None
            if isinstance(v, bool):
                return None
            if isinstance(v, (int, float)):
                return float(v)
            s = str(v).strip()
            neg = s.startswith('(') and s.endswith(')')
            s2 = s.replace('¥', '').replace('$', '').replace(',', '').replace('(', '-').replace(')', '')
            try:
                return float(s2)
            except Exception:
                return None

        def _date(v):
            if v is None or (isinstance(float, type) and v != v):
                return None
            if hasattr(v, 'year'):
                return v.strftime('%Y-%m-%d')
            s = str(v).strip()
            m = re.search(r'(\d{4})[-/](\d{1,2})[-/](\d{1,2})', s)
            if m:
                return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
            return None

        # pandas: 列号 0-based → A=0, B=1, ..., T=19
        groups = []
        cur = None
        for idx in range(1, len(df)):  # 跳过表头行(row 0)
            row = df.iloc[idx]
            A = row.iloc[0] if len(row) > 0 else None   # 序号
            B = row.iloc[1] if len(row) > 1 else None   # 日期/状态词
            E = row.iloc[4] if len(row) > 4 else None   # 基金
            F = row.iloc[5] if len(row) > 5 else None   # 动作
            N = row.iloc[13] if len(row) > 13 else None  # 收益RMB(N列=14→idx13)
            O = row.iloc[14] if len(row) > 14 else None  # 对冲量(O列=15→idx14)
            P = row.iloc[15] if len(row) > 15 else None  # 对冲价(P列=16→idx15)
            S = row.iloc[18] if len(row) > 18 else None  # IB小计USD(S列=19→idx18)
            I = row.iloc[8] if len(row) > 8 else None    # 单价(I列=9→idx8)
            K = row.iloc[10] if len(row) > 10 else None  # 备注(K列=11→idx10)
            L = row.iloc[11] if len(row) > 11 else None  # 美股标的(L列=12→idx11)
            Q = row.iloc[16] if len(row) > 16 else None  # 对冲额(Q列=17→idx16)
            R = row.iloc[17] if len(row) > 17 else None  # 佣金(R列=18→idx17)

            # [AI-2026-08-18] 空行判定须含 G/O/S（同 openpyxl 版）：纯美股对冲行（F 空但 O/S 有值）不是空行
            Gv = row.iloc[6] if len(row) > 6 else None
            if (pd.isna(A) and pd.isna(B) and pd.isna(E) and pd.isna(F) and pd.isna(N)
                    and pd.isna(Gv) and pd.isna(O) and pd.isna(S)):
                # [AI-2026-08-18 东哥口径] 同 openpyxl 版：空行只是格式占位，组结束唯一标志=汇总行
                continue

            E = str(E).strip() if not pd.isna(E) else ''
            F = str(F).strip() if not pd.isna(F) else ''
            A_str = str(int(A)) if not pd.isna(A) and isinstance(A, (int, float)) else (str(A).strip() if not pd.isna(A) else '')
            B_str = str(B).strip() if not pd.isna(B) else ''

            if F == '汇总':
                if cur is None:
                    cur = {'details': [], 'fund': E}
                cur['fund'] = E or cur.get('fund')
                b_is_date = hasattr(B, 'year') or (B_str and B_str.lower() not in STATUS_WORDS)
                status = self._STATUS_MAP.get(B_str.lower(), 'Final' if b_is_date else B_str)
                G = row.iloc[6] if len(row) > 6 else None  # G列=多账户交易金额
                summary = {
                    'serial': A_str,
                    'status': status,
                    'fund': E,
                    'pnl_rmb': _num(N),
                    'pnl_usd': _num(S),
                    'buy_amount': _num(G),
                    'buy_price': _num(I),
                    'short_price': _num(P),
                    'short_amount': _num(Q),
                    'us_commission': _num(R),
                    'hedge_symbol': (str(L).strip() if not pd.isna(L) else ''),
                    'buy_notes': (str(K).strip() if not pd.isna(K) else ''),
                    'buy_date': None,
                    'sell_date': None,
                }
                cur['summary'] = summary
                groups.append(cur)
                cur = None
                continue

            has_us = (not pd.isna(N)) or (not pd.isna(O)) or (not pd.isna(S))
            if E or has_us:
                if cur is None:
                    cur = {'details': [], 'fund': E}
                if E:
                    cur['fund'] = E
                J_val = row.iloc[9] if len(row) > 9 else None  # J列=数量(idx9)
                I_val = row.iloc[8] if len(row) > 8 else None  # I列=单价(idx8)
                R_val = row.iloc[17] if len(row) > 17 else None  # R列=佣金(idx17)
                G_val = row.iloc[6] if len(row) > 6 else None  # G列 A股交易金额
                S_val = row.iloc[18] if len(row) > 18 else None  # S列 IB小计USD
                cur['details'].append({
                    'action': F or '美股明细',
                    'date': _date(B),
                    'volume': _num(J_val),
                    'short_vol': _num(O),
                    'short_price': _num(P),
                    'price': _num(I_val),
                    'comm': _num(R_val),
                    'g_amt': _num(G_val),
                    's_amt': _num(S_val),
                })
        if cur:
            if 'summary' not in cur:
                logger.warning(f"V7 第 {idx + 1} 行附近（文件末尾）：存在未以'汇总'行结尾的明细块，该组已跳过（不影响其他组）")
            groups.append(cur)
        return groups

    def import_v7(self, file_path: str) -> Dict[str, Any]:
        """解析 v7 Excel 并 upsert 到 arbitrage_pairs。
        匹配键：serial_no（已存在则 UPDATE，不存在则 INSERT）。V7 即唯一真源，导入=把 DB 同步到 V7。
        盈亏口径（第一性原理，全部程序从明细重算，不读汇总行手填的 G/N/S——旧数据手填值多错）：
          - A股盈亏 a_share_pnl = 该组所有明细行 G 列求和（买入负/赎回正，自然得净盈亏）
          - 美股盈亏 pnl_usd     = 该组所有明细行 S 列求和（卖空正/买平负，自然得净盈亏）
          - 总盈亏   pnl_rmb     = a_share_pnl + pnl_usd × 平仓日汇率
          - 仅 Closed 落库；OPEN/unfinished 一律 NULL。
        金额归一：V7 G 列(LOF买入总额)习惯写负数，OPEN/unfinished 行取 abs 存正数(与历史一致)；
        Closed 行 G 为净额，不写入 buy_amount。
        DB 中存在但 V7 不存在的序号（孤儿行）一律保留不动。
        返回 {inserted, updated, skipped, errors}。
        """
        groups = self._parse_v7_groups(file_path)
        conn = self.db._get_conn()
        try:
            self._ensure_a_share_pnl(conn)
            rates = self._get_rates_map(conn)
            # [AI-2026-08-18] fund_name 从主库 unified_fund_list 查（V7 只有基金代码没有名字），
            # 否则新组 INSERT 后 fund_name=NULL → 前端"即将赎回"卡片显示 null
            fund_names = {}
            try:
                m = self._attach_master(conn)
                for r in conn.execute(
                        f"SELECT fund_code, fund_name FROM {m}.unified_fund_list WHERE fund_name IS NOT NULL").fetchall():
                    fund_names[str(r[0])] = r[1]
            except Exception as e:
                logger.warning(f"[AI-2026-08-18] 读取 fund_name 映射失败: {e}")
            inserted = 0
            updated = 0
            skipped = 0
            errors = []
            for g in groups:
                s = g.get('summary')
                if not s:
                    logger.warning(f"V7 解析到一组缺少'汇总'行的明细块（fund={g.get('fund')}），已跳过")
                    skipped += 1
                    continue
                fund = (s.get('fund') or '').strip()
                if not fund:
                    errors.append(f"组缺少基金代码 (pnl_rmb={s.get('pnl_rmb')})")
                    skipped += 1
                    continue
                fund_name = self._HABIT_FUND_NAMES.get(fund) or fund_names.get(fund, fund)  # 习惯叫法优先，缺则官方名/代码
                status = s.get('status') or 'Closed'
                if status == 'Final':   # 历史误把日期当状态词 -> 归一为 Closed
                    status = 'Closed'
                buys = [d for d in g['details'] if d['action'] in ('买入', '开仓续')]
                sells = [d for d in g['details'] if d['action'] in ('卖出', '赎回')]
                buy_date = s.get('buy_date') or (buys[0]['date'] if buys else None)
                sell_date = s.get('sell_date') or (sells[-1]['date'] if sells else None)
                # [AI-2026-08-18] 盈亏从明细求和重算（第一性原理，不读汇总行手填值）
                # A股盈亏 = Σ 明细 G（买入负/赎回正）；美股盈亏 = Σ 明细 S（卖空正/买平负）
                a_sum = round(sum((d.get('g_amt') or 0) for d in g['details']), 2) or None
                u_sum = round(sum((d.get('s_amt') or 0) for d in g['details']), 2) or None
                if status == 'Closed':
                    a_share_pnl = a_sum
                    pnl_usd = u_sum
                    # [AI-2026-08-18] 汇率按 平仓日→开仓日 取主库中间价；无买卖腿（瘸腿组）取明细最早日期；缺失不兜底
                    ref_date = sell_date or buy_date
                    if not ref_date:
                        _dates = sorted(d['date'] for d in g['details'] if d.get('date'))
                        ref_date = _dates[0] if _dates else None
                    rate = self._rate_on(rates, ref_date)
                    if rate is None:
                        pnl_rmb = None
                        logger.warning(
                            f"[AI-2026-08-18] 组 {s.get('serial')} 缺汇率：ref_date={ref_date}，pnl_rmb 留空，请补日期后重导")
                    elif pnl_usd is None:
                        pnl_rmb = round(a_share_pnl, 2) if a_share_pnl is not None else None
                    else:
                        # 瘸腿组（无 LOF 腿 a_share_pnl=None）美股单独存在也能算总盈亏
                        pnl_rmb = round((a_share_pnl or 0) + pnl_usd * rate, 2)
                else:
                    a_share_pnl = None
                    pnl_usd = None
                    pnl_rmb = None
                # [AI-2026-08-18] 数量保留 V7 真实符号（买正/卖负/空负/平正），不再 abs
                buy_volume = sum(d['volume'] or 0 for d in buys) or None
                sell_volume = sum(d['volume'] or 0 for d in sells) or None
                short_volume = sum(d['short_vol'] or 0 for d in g['details'] if (d.get('short_vol') or 0) < 0) or None
                cover_volume = sum(d['short_vol'] or 0 for d in g['details'] if (d.get('short_vol') or 0) > 0) or None
                hedge = s.get('hedge_symbol') or self._HEDGE_MAP.get(fund, '')
                broker = '华宝' if fund == '162411' else '银河'
                # [AI-2026-08-18] 东哥口径：价格 = 金额求和 ÷ 数量求和（不读汇总手填 I/P 列）
                # 开仓价 = Σ|G负| ÷ ΣJ（买入行）；平仓价 = ΣG正 ÷ Σ|J|（赎回行）
                # 空单价 = ΣS正 ÷ Σ|O|（卖空开仓腿）；买平价 = Σ|S负| ÷ ΣO（买平腿）
                # 金额也全部从明细求和：buy_amount=Σ|G负|, sell_amount=ΣG正,
                # short_amount=ΣS正, cover_amount=Σ|S负|, us_commission=Σ|R|
                def _sum_by(rows, key):
                    return sum((d.get(key) or 0) for d in rows) or None
                buy_rows = [d for d in buys if (d.get('g_amt') or 0) < 0]
                sell_rows = [d for d in sells if (d.get('g_amt') or 0) > 0]
                short_rows = [d for d in g['details'] if (d.get('s_amt') or 0) > 0]
                cover_rows = [d for d in g['details'] if (d.get('s_amt') or 0) < 0]
                # 开仓价 = Σ|G负| / ΣJ
                buy_g = sum(abs(d.get('g_amt') or 0) for d in buy_rows)
                buy_j = sum(abs(d.get('volume') or 0) for d in buy_rows)
                buy_price = round(buy_g / buy_j, 6) if buy_j else None
                # 平仓价 = ΣG正 / Σ|J|
                sell_g = sum(abs(d.get('g_amt') or 0) for d in sell_rows)
                sell_j = sum(abs(d.get('volume') or 0) for d in sell_rows)
                sell_price = round(sell_g / sell_j, 6) if sell_j else None
                # 空单价 = ΣS正 / Σ|O|（卖空腿 O<0）
                short_s = sum(abs(d.get('s_amt') or 0) for d in short_rows)
                short_o = sum(abs(d.get('short_vol') or 0) for d in short_rows if (d.get('short_vol') or 0) < 0)
                short_price = round(short_s / short_o, 6) if short_o else None
                # 买平价 = Σ|S负| / ΣO（买平腿 O>0）
                cover_s = sum(abs(d.get('s_amt') or 0) for d in cover_rows)
                cover_o = sum(abs(d.get('short_vol') or 0) for d in cover_rows if (d.get('short_vol') or 0) > 0)
                cover_price = round(cover_s / cover_o, 6) if cover_o else None
                # [AI-2026-08-18] 金额/佣金保留 V7 真实符号（东哥否决 abs）：买负卖正/卖空正/买平负/佣金负
                buy_amount = round(-buy_g, 2) if buy_g else None        # 买入支出 → 负
                sell_amount = round(sell_g, 2) if sell_g else None      # 卖出/赎回收入 → 正
                short_amount = round(short_s, 2) if short_s else None   # 卖空收入 → 正
                cover_amount = round(-cover_s, 2) if cover_s else None  # 买平支出 → 负
                us_commission = round(sum(d.get('comm') or 0 for d in g['details']), 2) or None  # R 原样（V7 通常负）
                short_date = buy_date
                # [AI-2026-08-18] open/close_type 按 F 词表真实值（原硬编码 BUY/REDEEM，场内卖出被误标为赎回）
                open_type = 'ADD' if any(d.get('action') == '开仓续' for d in buys) else 'BUY'
                close_type = 'SELL' if any(d.get('action') == '卖出' for d in sells) else 'REDEEM'
                buy_notes = s.get('buy_notes') or ''

                serial = s.get('serial')
                if not serial:
                    errors.append(f"组缺少序号 (fund={fund})")
                    skipped += 1
                    continue
                # [2026-08-18] upsert: 以 serial_no 为主键, 已存在则 UPDATE(同步 V7), 不存在则 INSERT
                now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                existing = conn.execute(
                    "SELECT id FROM arbitrage_pairs WHERE serial_no=?", (serial,)
                ).fetchone()
                if existing:
                    conn.execute(
                        """UPDATE arbitrage_pairs SET
                            fund_code=?, fund_name=?, buy_date=?, sell_date=?, buy_volume=?, sell_volume=?, short_volume=?, cover_volume=?,
                            buy_amount=?, buy_price=?, sell_amount=?, sell_price=?, short_date=?, short_price=?, short_amount=?,
                            cover_date=?, cover_price=?, cover_amount=?, us_commission=?,
                            pnl_rmb=?, pnl_usd=?, a_share_pnl=?, status=?, hedge_symbol=?, broker_name=?,
                            buy_notes=?, open_type=?, close_type=?, updated_at=?
                            WHERE id=?""",
                        (fund, fund_name, buy_date, sell_date, buy_volume, sell_volume, short_volume, cover_volume,
                         buy_amount, buy_price, sell_amount, sell_price, short_date, short_price, short_amount,
                         sell_date, cover_price, cover_amount, us_commission,
                         pnl_rmb, pnl_usd, a_share_pnl, status, hedge, broker,
                         buy_notes, open_type, close_type, now, existing[0])
                    )
                    updated += 1
                else:
                    conn.execute(
                        """INSERT INTO arbitrage_pairs
                            (fund_code, fund_name, buy_date, sell_date, buy_volume, sell_volume, short_volume, cover_volume,
                             buy_amount, buy_price, sell_amount, sell_price, short_date, short_price, short_amount,
                             cover_date, cover_price, cover_amount, us_commission,
                             pnl_rmb, pnl_usd, a_share_pnl, status, hedge_symbol, broker_name, serial_no,
                             buy_notes, open_type, close_type, created_at, updated_at)
                            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                        (fund, fund_name, buy_date, sell_date, buy_volume, sell_volume, short_volume, cover_volume,
                         buy_amount, buy_price, sell_amount, sell_price, short_date, short_price, short_amount,
                         sell_date, cover_price, cover_amount, us_commission,
                         pnl_rmb, pnl_usd, a_share_pnl, status, hedge, broker, serial,
                         buy_notes, open_type, close_type, now, now)
                    )
                    inserted += 1
            conn.commit()
            # [AI-2026-08-18] 汇率回填 V7 汇总行 T 列 + N 列公式（用户 V7 为原始账本/核对基准，免手查汇率）
            try:
                self._write_back_v7(file_path, rates, groups)
            except Exception as e:
                logger.error(f"[AI-2026-08-18] 写回 V7（T列汇率/N列公式）失败: {e}")
            return {'inserted': inserted, 'updated': updated, 'skipped': skipped, 'errors': errors}
        except Exception as e:
            logger.error(f"导入 v7 失败: {e}")
            raise
        finally:
            conn.close()

    def _write_back_v7(self, file_path: str, rates: dict, groups: list) -> int:
        """[AI-2026-08-18] 导入后回填 V7 Excel：汇总行 T 列写汇率（主库按平仓日→开仓日取），
        N 列写公式 =G+S*T 覆盖旧手填值。用户 V7 是原始账本+核对基准：
        T 程序回填免手查汇率，N 由 Excel 公式自动算收益，与库 pnl_rmb 同源可核对。
        仅回填 Closed 组；非 Closed（OPEN/unfinished）不动（无最终盈亏）。
        """
        import openpyxl
        wb = openpyxl.load_workbook(file_path)  # data_only=False：保留公式
        ws = wb.active
        written = 0
        for g in groups:
            s = g.get('summary') or {}
            row = s.get('row')
            if not row or s.get('status') != 'Closed':
                continue
            buys = [d for d in g['details'] if d['action'] in ('买入', '开仓续')]
            sells = [d for d in g['details'] if d['action'] in ('卖出', '赎回')]
            buy_date = s.get('buy_date') or (buys[0]['date'] if buys else None)
            sell_date = s.get('sell_date') or (sells[-1]['date'] if sells else None)
            ref_date = sell_date or buy_date
            if not ref_date:
                _dates = sorted(d['date'] for d in g['details'] if d.get('date'))
                ref_date = _dates[0] if _dates else None
            rate = self._rate_on(rates, ref_date)
            if rate is None:
                continue  # 缺汇率：import 已 WARNING，此处不写，避免污染
            ws.cell(row, 20).value = rate                 # T 列 = 汇率
            ws.cell(row, 14).value = f"=G{row}+S{row}*T{row}"  # N 列 = 收益公式
            written += 1
        if written:
            wb.save(file_path)
            logger.info(f"[AI-2026-08-18] V7 回填完成：{written} 个 Closed 组写入汇率+收益公式")
        return written

    def get_ledger_alerts(self) -> Dict[str, Any]:
        """赎回提醒：OPEN 笔推算可优惠赎回日/倒计时/告警级别；unfinished 提示待净值。"""
        pairs = self.get_all_pairs()
        today = datetime.now().date()
        _wd_cn = ['周一', '周二', '周三', '周四', '周五', '周六', '周日']
        open_alerts = []
        unfinished_alerts = []
        for p in pairs:
            status = (p.get('status') or '').strip()
            if status == 'OPEN':
                buy_date = p.get('buy_date')
                if not buy_date:
                    continue
                redeem_d = self._discount_redeem_date(buy_date)
                if redeem_d is None:
                    continue
                days_left = (redeem_d - today).days
                wd_cn = _wd_cn[redeem_d.weekday()]
                # [AI-2026-08-20] 只显示下一个交易日需要赎回的（days_left <= 1）
                if days_left > 1:
                    continue
                if days_left <= 0:
                    level, msg = 'critical', f"已到可优惠赎回日（{redeem_d.isoformat()} {wd_cn}），请赎回"
                else:
                    level, msg = 'warning', f"明天（{redeem_d.isoformat()} {wd_cn}）即可优惠赎回"
                open_alerts.append({
                    'id': p.get('id'), 'fund_code': p.get('fund_code'), 'fund_name': p.get('fund_name'),
                    'buy_date': buy_date, 'redeem_date': redeem_d.isoformat(), 'redeem_weekday': wd_cn,
                    'days_left': days_left, 'level': level, 'message': msg,
                    'short_volume': p.get('short_volume'), 'short_price': p.get('short_price'),
                    'buy_volume': p.get('buy_volume'),
                    'pnl_rmb': p.get('pnl_rmb'), 'pnl_usd': p.get('pnl_usd'),
                })
            elif status == 'unfinished':
                unfinished_alerts.append({
                    'id': p.get('id'), 'fund_code': p.get('fund_code'), 'fund_name': p.get('fund_name'),
                    'buy_date': p.get('buy_date'), 'sell_date': p.get('sell_date'),
                    'level': 'warning', 'message': '已有退出动作，待净值/结算',
                })
        open_alerts.sort(key=lambda x: (x['days_left'] if x['days_left'] is not None else 999))
        return {
            'today': today.isoformat(),
            'open_count': len(open_alerts),
            'unfinished_count': len(unfinished_alerts),
            'open_alerts': open_alerts,
            'unfinished_alerts': unfinished_alerts,
        }

